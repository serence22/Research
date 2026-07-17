from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from phosphorus_sublimation_core import (
    DATA_FIRST_ROW_EXCEL,
    DATA_LAST_ROW_EXCEL,
    HEADER_ROW_EXCEL,
    SHEET_NAME,
    analyze_file,
    build_records,
    fit_arrhenius,
    parse_pressure_torr,
    parse_temperature_raw,
    read_source_sheet,
    summarize_by_temperature,
)


ROOT = Path(__file__).resolve().parents[1]
SOURCE = next(ROOT.glob("*.xlsx"))


HEADERS = [
    "CVD 종류",
    "공정 압력",
    "온도",
    "T1(min.)",
    "T2(min.)",
    "p1 (mg)",
    "p2(mg)",
    "T2-T1",
    "분당 소모율(%/min)",
    "10분당 소모율(%/10min)",
]


def make_workbook(rows: list[list[object]], extra_row: list[object] | None = None, other_sheet_bad: bool = False) -> Path:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=HEADER_ROW_EXCEL, column=col, value=header)
    for offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            ws.cell(row=DATA_FIRST_ROW_EXCEL + offset, column=col, value=value)
    if extra_row:
        for col, value in enumerate(extra_row, start=1):
            ws.cell(row=DATA_LAST_ROW_EXCEL + 1, column=col, value=value)
    if other_sheet_bad:
        other = wb.create_sheet("다른시트")
        for col, header in enumerate(HEADERS, start=1):
            other.cell(row=HEADER_ROW_EXCEL, column=col, value=header)
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=1, value="SHOULD_NOT_APPEAR")
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=2, value="0.15 torr")
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=3, value=999)
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=4, value=30)
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=5, value=40)
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=6, value=200)
        other.cell(row=DATA_FIRST_ROW_EXCEL, column=7, value=100)
    wb.save(tmp.name)
    return Path(tmp.name)


class PhosphorusSublimationTests(unittest.TestCase):
    def test_reads_formula_sheet(self) -> None:
        df = read_source_sheet(SOURCE)
        self.assertFalse(df.empty)

    def test_header_row_90_is_used(self) -> None:
        df = read_source_sheet(SOURCE)
        self.assertEqual(list(df.columns[:3]), ["CVD 종류", "공정 압력", "온도"])

    def test_data_rows_91_to_106_only_are_read(self) -> None:
        df = read_source_sheet(SOURCE)
        self.assertEqual(len(df), 16)

    def test_missing_formula_sheet_has_clear_error(self) -> None:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb = Workbook()
        wb.active.title = "다른시트"
        wb.save(tmp.name)
        with self.assertRaisesRegex(ValueError, "수식"):
            read_source_sheet(tmp.name)

    def test_other_sheet_data_is_not_included(self) -> None:
        path = make_workbook(
            [["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 999, 999]],
            other_sheet_bad=True,
        )
        result = analyze_file(path)
        self.assertTrue(all(r.cvd_type != "SHOULD_NOT_APPEAR" for r in result.records))
        self.assertTrue(all(r.temperature_c != 999 for r in result.records))

    def test_ignores_rows_after_106(self) -> None:
        path = make_workbook(
            [["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 0, 0]],
            extra_row=["CVD 1", "0.15 torr", 999, 30, 40, 200, 100, 10, 0, 0],
        )
        result = analyze_file(path)
        temps = [r.temperature_c for r in result.records if r.temperature_c is not None]
        self.assertNotIn(999, temps)

    def test_pressure_parsing(self) -> None:
        self.assertAlmostEqual(parse_pressure_torr("0.15 torr"), 0.15)

    def test_0029_torr_excluded_from_fitting(self) -> None:
        result = analyze_file(SOURCE)
        self.assertEqual(sum(1 for r in result.records if r.included), 9)
        self.assertTrue(all(abs(r.pressure_torr - 0.15) <= 0.005 for r in result.records if r.included))
        self.assertEqual(result.fit.interval_count, 9)

    def test_temperature_parsing_and_forward_fill(self) -> None:
        self.assertEqual(parse_temperature_raw("365(MFC)=374"), (374.0, True))
        result = analyze_file(SOURCE)
        row98 = next(r for r in result.records if r.excel_row == 98)
        row99 = next(r for r in result.records if r.excel_row == 99)
        self.assertEqual(row98.temperature_c, 374)
        self.assertEqual(row99.temperature_c, 374)

    def test_interval_rate_calculation(self) -> None:
        path = make_workbook([["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 999, 999]])
        record = analyze_file(path).records[0]
        self.assertAlmostEqual(record.interval_loss_mg, 14.1)
        self.assertAlmostEqual(record.rate_mg_per_min, 1.41)
        self.assertAlmostEqual(record.rate_mg_per_hour, 84.6)

    def test_percent_columns_do_not_change_mg_per_min(self) -> None:
        rows_a = [["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 0.01, 9.6]]
        rows_b = [["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 12345, 67890]]
        rate_a = analyze_file(make_workbook(rows_a)).records[0].rate_mg_per_min
        rate_b = analyze_file(make_workbook(rows_b)).records[0].rate_mg_per_min
        self.assertEqual(rate_a, rate_b)

    def test_no_moisture_or_200mg_denominator_correction(self) -> None:
        path = make_workbook([["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 0, 0]])
        record = analyze_file(path).records[0]
        self.assertAlmostEqual(record.rate_mg_per_min, (146.8 - 132.7) / (40 - 30))

    def test_fitting_stops_with_less_than_three_temperatures(self) -> None:
        rows = [
            ["CVD 1", "0.15 torr", 390, 30, 40, 146.8, 132.7, 10, 0, 0],
            ["CVD 1", "0.15 torr", 375, 30, 40, 176.8, 171.5, 10, 0, 0],
        ]
        records = build_records(read_source_sheet(make_workbook(rows)))
        summaries = summarize_by_temperature(records, "Duration-weighted mean")
        self.assertIsNone(fit_arrhenius(summaries))


if __name__ == "__main__":
    unittest.main()
